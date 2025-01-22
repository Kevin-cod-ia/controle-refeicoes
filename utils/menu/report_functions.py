from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse
from menu.models import Employee, Company, WeekMenu, Unity, UserChoice
from openpyxl.utils import get_column_letter
import locale

def generate_unity_options_report(wb, unit_name, unit_address, unit_filter):
    total_unit_employees_day = 0
    # Cria uma nova aba para a unidade
    ws = wb.create_sheet(title=f"OPÇÕES {unit_name.upper()}")

    # Definindo a largura das colunas manualmente antes de adicionar dados
    column_widths = {
        1: 40,  # Coluna 1 (Prato Principal)
        2: 17,  # Coluna 2 (Total de Opções)
        3: 25,  # Coluna 3 (Omelete)
        4: 30,  # Coluna 4 (Marmita Fit Frango)
        5: 30,  # Coluna 5 (Marmita Fit Carne)
        6: 20,  # Coluna 6 (Marmita Fit Vegana)
    }
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Título principal
    title = f"{unit_name.upper()} - {unit_address}"
    ws.merge_cells("A2:F2")
    title_cell = ws["A2"]
    title_cell.value = title
    title_cell.font = Font(name="Verdana", size=14, bold=True, underline="single")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    title_cell.fill = PatternFill(start_color="B1B0B6", end_color="B1B0B6", fill_type="solid")  # Fundo cinza claro

    # Ajuste o tamanho da linha para adicionar espaçamento superior e inferior
    ws.row_dimensions[2].height = 30

    # Definindo a altura das linhas manualmente antes de adicionar dados
    row_heights = {
        3: 70,  # Linha 3 (Cabeçalho da Tabela)
        4: 55,  # Linha 4 (Coluna Segunda-feira)
        5: 55,  # Linha 5 (Coluna Terça-feira)
        6: 55,  # Linha 6 (Coluna Quarta-feira)
        7: 55,  # Linha 7 (Coluna Quinta-feira)
        8: 55,  # Linha 8 (Coluna Sexta-feira)
        9: 55,  # Linha 9 (Linha extra)

    }
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height

    # Cabeçalhos da tabela
    headers = [
        "Prato Principal",
        "Total de Opções",
        "Omelete",
        "Opções - Marmita Fit Frango",
        "Opções - Marmita Fit Carne",
        "Marmita Fit Vegana",
    ]
    ws.append(headers)

    # Definir a formatação do cabeçalho da tabela (linha 3)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Verdana", size=12, bold=True, underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # Preenchimento da tabela
    week_menu = WeekMenu.objects.filter()
    row = 4


    for menu in week_menu:
        total_options = UserChoice.objects.filter(
            menu=menu, user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        ).count()
        omelet_count = UserChoice.objects.filter(
            menu=menu, option__name_option="Omelete", user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        ).count()
        chicken_count = UserChoice.objects.filter(
            menu=menu, option__name_option="Marmita de Frango", user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        ).count()
        beef_count = UserChoice.objects.filter(
            menu=menu, option__name_option="Marmita de Carne", user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        ).count()


        # Pega todos os funcionários que pediram determinada opção
        omelet_employees = UserChoice.objects.filter(
            menu=menu, option__name_option="Omelete", user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        )
        chicken_employees = UserChoice.objects.filter(
            menu=menu, option__name_option="Marmita de Frango", user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        )
        beef_employees = UserChoice.objects.filter(
            menu=menu, option__name_option="Marmita de Carne", user__employee__unity__unity_name=unit_filter, user__employee__is_on_vacations=False
        )

        # Gera uma lista de nomes de funcionários que pediram determinada opção
        omelet_names = [f'{user_choice.user.employee.first_name} {user_choice.user.employee.last_name}' for user_choice in omelet_employees]
        chicken_names = [f'{user_choice.user.employee.first_name} {user_choice.user.employee.last_name}' for user_choice in chicken_employees]
        beef_names = [f'{user_choice.user.employee.first_name} {user_choice.user.employee.last_name}' for user_choice in beef_employees]

        # Cria a string formatada com os nomes dos funcionários
        omelet_names_str = " / ".join(omelet_names)
        chicken_names_str = " / ".join(chicken_names)
        beef_names_str = " / ".join(beef_names)

        # Adiciona os dados na tabela

        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil')
        week_day = menu.date_meal.strftime("%A")
    
        ws.append([
            f'{menu.date_meal.strftime("%d.%m")} ({week_day.upper()}) - {menu.title.upper()}',
            total_options,
            f'{omelet_count} - {omelet_names_str}' if omelet_count > 0 else omelet_count,
            f'{chicken_count} - {chicken_names_str}' if chicken_count > 0 else chicken_count,
            f'{beef_count} - {beef_names_str}' if beef_count > 0 else beef_count,
            "",  # Coluna F permanece vazia
        ])
        

        # Formatação das células
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Verdana", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            if col == 2:  # Coluna B (Total de Opções) em vermelho
                cell.font = Font(name="Verdana", size=12, color="FF0000", bold=True)
        row += 1

    row += 1
    ws.append(["",  "",  ""  ])
    for menu in week_menu:
    # Adiciona título "Faturamento do determinado dia"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        title_cell = ws.cell(row=row, column=1, value=f'FATURAMENTO {menu.date_meal.strftime("%d.%m")} ({menu.date_meal.strftime("%A").upper()})')
        title_cell.font = Font(name="Verdana", size=15, bold=True, underline="single")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row += 2  # Pula uma linha em branco para dar espaçamento


        # Adiciona a tabela de faturamento
        ws.append(["",  "",  ""  ])

        options_meal_list = ['ALMOÇO', 'OMELETE', 'MARMITAS FIT']

        total_unit_employees_day = 0

        for i, option in enumerate(options_meal_list):
            active_employees = Employee.objects.filter(unity__unity_name=unit_name, is_on_vacations=False)
            
            
            if option == 'ALMOÇO':
                # Funcionários que não pediram nem omelete nem marmita
                no_order_count = active_employees.exclude(
                    user__userchoice__menu=menu,
                    user__userchoice__option__name_option__in=["Omelete", "Marmita de Frango", "Marmita de Carne"]
                ).count()
                total_value = no_order_count
            elif option == 'OMELETE':
                # Funcionários que pediram omelete
                omelet_count = UserChoice.objects.filter(
                    menu=menu,
                    option__name_option="Omelete",
                    user__employee__unity__unity_name=unit_name,
                    user__employee__is_on_vacations=False
                    
                ).count()
                total_value = omelet_count
            elif option == 'MARMITAS FIT':
                # Funcionários que pediram marmita de frango ou carne
                marmita_count = UserChoice.objects.filter(
                    menu=menu,
                    option__name_option__in=["Marmita de Frango", "Marmita de Carne"],
                    user__employee__unity__unity_name=unit_name,
                    user__employee__is_on_vacations=False
                ).count()
                total_value = marmita_count

            # Adiciona a linha de faturamento
            ws.append([
                f"  {unit_name.upper()} - {unit_address} ",  # Coluna A
                f"{option}",  # Coluna B
                "",  # Coluna C
                f"  {total_value}"  # Coluna D
            ])
  
            total_unit_employees_day = total_unit_employees_day + total_value

            # Formatação da linha adicionada
            for col in range(1, 5):  # Colunas A a D
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style="thick"),
                    right=Side(style="thick"),
                    top=Side(style="thick"),
                    bottom=Side(style="thick"),
                )
                if col == 1:  # Coluna A (Unidade)
                    cell.font = Font(name="Verdana", size=10, bold=True, underline="single")
                elif col == 2:  # Coluna B (Opções de refeição)
                    cell.font = Font(name="Verdana", size=10)
                elif col == 4:  # Coluna D (Total de funcionários)
                    cell.font = Font(name="Verdana", size=14, bold=True, color="FF0000")
            

            row += 3  # Pula uma linha em branco para dar espaçamento
            ws.append(["",  "",  ""  ])
            ws.append(["",  "",  ""  ])

        ws.append(["", "Total", "",  f"{total_unit_employees_day}"  ])   

        # Formatação da linha do total

        for col in range(2, 5):  # Colunas B a D
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Verdana", size=16, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 4:  # Coluna D (Valor total)
                cell.font = Font(name="Verdana", size=16, bold=True)  # Verde
        # Incrementa a linha para novos dados (se necessário)
        row += 2




def generate_full_report_function(request):
    wb = Workbook()
    ws = wb.active
    week_menu = WeekMenu.objects.filter()
    first_day = week_menu[0].date_meal.strftime("%d.%m.%Y")
    last_day = week_menu[4].date_meal.strftime("%d.%m.%Y")
    ws.title = f"{first_day} a {last_day}"

    # Adiciona cabeçalho da tabela de funcionários
    headers = ['Unidade', 'Turno', 'Nome da Empresa', 'Nome do Funcionário']
    ws.append([f"  {header}  " for header in headers])  # Adiciona espaçamento horizontal

    # Aplica a formatação do cabeçalho
    for cell in ws[1]:
        cell.font = Font(name="Verdana", size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Mapeia os nomes completos das empresas
    COMPANY_FULL_NAMES = {
        'FG&P': 'FG&P CONSULTORIA ADMINISTRATIVA LTDA.',
        'FCD': 'FCD ARMAZENAGEM E DISTRIBUIÇÃO LTDA',
        'Sustenpack': 'SUSTENPACK EMBALAGENS SUSTENTAVEIS IMPORTACAO E EX',
    }

    row = 2  # Começa a adicionar os dados a partir da segunda linha

    # Agrupar os funcionários por empresa
    employees = Employee.objects.select_related('company', 'unity', 'shift').order_by('company__company_name', 'first_name', 'last_name')
    employees_by_company = {}

    for employee in employees:
        company_name = employee.company.company_name
        full_company_name = COMPANY_FULL_NAMES.get(company_name, company_name)
        if full_company_name not in employees_by_company:
            employees_by_company[full_company_name] = []
        employees_by_company[full_company_name].append(employee)

    # Adiciona os funcionários à planilha
    for full_company_name, employees in employees_by_company.items():
        total_employees = len(employees)

        for employee in employees:
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") if employee.is_on_vacations else None

            # Adiciona os dados do funcionário na planilha
            ws.append([ 
                f"  {employee.unity.unity_name.upper()}  ",
                f"  {employee.shift.shift.upper()}  ",
                f"  {full_company_name.upper()}  ",
                f"  {employee.__str__().upper()}  "
            ])

            # Aplica estilo às células da linha
            for cell in ws[row]:
                cell.font = Font(name="Verdana", size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if fill:
                    cell.fill = fill

            row += 1

        # Adiciona a linha de total de funcionários
        ws.append(['', '', f"  {full_company_name.upper()}  ", f"  TOTAL: {total_employees} FUNCIONÁRIOS  "])
        for cell in ws[row]:
            cell.font = Font(name="Verdana", size=10, bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        row += 1

    # Aplica o filtro na tabela
    filter_range = f"A1:D{row - 1}"
    ws.auto_filter.ref = filter_range

    # Adiciona espaço antes da seção de faturamento
    row += 9

    # Adiciona título "Faturamento"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    title_cell = ws.cell(row=row, column=1, value=f"FATURAMENTO  {first_day} A {last_day}")
    title_cell.font = Font(name="Verdana", size=15, bold=True, underline="single")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2  # Pula uma linha em branco para dar espaçamento

    # Adiciona a tabela de faturamento
    ws.append(["",  "",  ""  ])
    total_general = 0
    for company_name, employees in employees_by_company.items():
        active_employees = [e for e in employees if not e.is_on_vacations]
        total_active = len(active_employees)
        total_general += total_active

        ws.append([
            f"  {company_name.upper()}  ",  # Coluna A
            "",  # Coluna B (vazia)
            f"  {total_active}  "  # Coluna C
        ])

        # Estiliza as células da tabela
        for col_index in range(1, 4):  # Apenas até a coluna C
            cell = ws.cell(row=row, column=col_index)
            cell.font = Font(name="Verdana", size=11)
            cell.border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    # Adiciona total geral
    ws.append(["", "", f"{total_general}"])
    total_cell = ws.cell(row=row, column=3)
    total_cell.font = Font(name="Verdana", size=14, bold=True)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")


    # Adiciona espaço antes da seção de faturamento (entrega)
    row += 2

    # Adiciona título "FATURAMENTO (ENTREGA)"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    delivery_title = ws.cell(row=row, column=1, value=f"FATURAMENTO (ENTREGA) {first_day} A {last_day}")
    delivery_title.font = Font(name="Verdana", size=15, bold=True, underline="single")
    delivery_title.alignment = Alignment(horizontal="center", vertical="center")
    row += 2  # Pula uma linha em branco para dar espaçamento

    # Adiciona a tabela de faturamento (entrega)
    ws.append(["",  "",  ""  ])
    for unity in Unity.objects.all():
        active_employees = Employee.objects.filter(unity=unity, is_on_vacations=False)
        total_active = len(active_employees)

        # Preenche a linha da unidade
        if unity.unity_name == "Unidade 1":
            unity_name = "UNIDADE 1 - Rua João Jose dos Reis, 59"
        elif unity.unity_name == "Unidade 2":
            unity_name = "UNIDADE 2 - Rua Jose Maria de Melo, 311"
        elif unity.unity_name == "Unidade 5 - Galpão Novo":
            unity_name = "UNIDADE 5 (GALPÃO NOVO) - Rua Jose Maria de Melo, 157"

        # Adiciona a linha de faturamento
        ws.append([
            f"  {unity_name}  ",  # Coluna A
            "ALMOÇO (CUBA)",  # Coluna B
            f"  {total_active}  "  # Coluna C
        ])

        # Estiliza as células da tabela
        for col_index in range(1, 4):  # Apenas até a coluna C
            cell = ws.cell(row=row, column=col_index)
            cell.font = Font(name="Verdana", size=11)
            cell.border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Aplica negrito e sublinhado apenas à parte "UNIDADE X"
            unit_title_cell = ws.cell(row=row, column=1)
            unit_title_cell.font = Font(name="Verdana", size=11, bold=True, underline="single")

            # Aplica a formatação vermelha e em negrito na coluna C
            cell_c = ws.cell(row=row, column=3)  # Coluna C
            cell_c.font = Font(name="Verdana", size=11, bold=True, color="FF0000")  # Vermelho e Negrito

        row += 2  # Pula uma linha para a próxima unidade
        ws.append(["",  "",  ""  ])

    # Adiciona a soma total geral de funcionários não de férias
    ws.append(["", "", f"  {total_general}  "])
    total_delivery_cell = ws.cell(row=row, column=3)
    total_delivery_cell.font = Font(name="Verdana", size=14, bold=True, color="FF0000")
    total_delivery_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 14
        ws.column_dimensions[column].width = adjusted_width



    # Gera a aba para a Unidade 1
    generate_unity_options_report(
        wb,
        unit_name="Unidade 1",
        unit_address="Rua João Jose dos Reis, 59",
        unit_filter="Unidade 1"
    )

    # Gera a aba para a Unidade 2
    generate_unity_options_report(
        wb,
        unit_name="Unidade 2",
        unit_address="Rua Jose Maria de Melo, 311",
        unit_filter="Unidade 2"
    )

    # Gera a aba para a Unidade 5 - Novo Galpão
    generate_unity_options_report(
        wb,
        unit_name="Unidade 5 - Novo Galpão",
        unit_address="Rua Jose Maria de Melo, 157",
        unit_filter="Unidade 5 - Novo Galpão"
    )
    

    # Retorna o arquivo como resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="RELACAO FUNCIONARIO JEITINHO BRASILEIRO {first_day} A {last_day}.xlsx"'

    wb.save(response)
    return response